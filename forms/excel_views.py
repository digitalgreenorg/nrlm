from openpyxl import Workbook
from openpyxl.style import Alignment, Border, Color, NumberFormat
from openpyxl.worksheet import RowDimension
from models import Project, Progress, ProgressTill13, State, Target

class MonthYear():
    def __init__(self, month, year):
        self.month = month
        self.year = year
        self.INDEX_TO_MONTH_MAP = {
            1 : "Jan",
            2: "Feb",
            3 : "Mar",
            4 : "Apr",
            5 : "May",
            6 : "Jun",
            7 : "Jul",
            8 : "Aug",
            9 : "Sep",
            10 : "Oct",
            11 : "Nov",
            12 : "Dec",
        }
        self.MONTH_TO_INDEX_MAP = {
            "Jan" : 1,
            "Feb" : 2,
            "Mar" : 3,
            "Apr" : 4,
            "May" : 5,
            "Jun" : 6,
            "Jul" : 7,
            "Aug" : 8,
            "Sep" : 9,
            "Oct" : 10,
            "Nov" : 11,
            "Dec" : 12
        }
        self.month_index = self.MONTH_TO_INDEX_MAP[self.month]
    
    def get_financial_year(self):
        if self.month_index not in [1,2,3]:
            return self.year + 1
        return self.year
    
    def is_financial_year_2013_14(self):
        if self.year >= 2014 and self.month_index not in [1,2,3]:
            return False
        return True
    
    def is_financial_year_2014_15(self):
        if self.year >= 2015 and self.month_index not in [1,2,3]:
            return False
        return True
        
    def get_previous_months(self):
        previous_months = {}
        if self.year == 2013:
            if self.month_index > 8:
                previous_months[2013] = self.get_month_range(9, self.month_index) 
            else:
                raise Exception
        elif self.is_financial_year_2013_14():
            previous_months[2013] = range(9,13)
            if month_index > 1:
                previous_months[2014] = self.get_month_range(1, self.month_index)
        elif self.month_index in [4,5,6,7,8,9,10,11,12]:
            previous_months[self.year] = self.get_month_range(4, self.month_index)
        else:
            previous_months[self.year-1] = self.get_month_range(4,13)
            if month > 1:
                previous_months[self.year] = self.get_month_range(1, self.month_index)
        return previous_months
    
    def get_all_months(self):
        return self.get_month_range(1,13)
    
    def get_previous_years(self):
        year = self.get_financial_year() - 1
        years = range(2014, year)
        return years
        
    def get_month_range(self, first_index, last_index):
        return [self.INDEX_TO_MONTH_MAP[x] for x in range(first_index, last_index)]

def get_project_excel_sheet(query_month, query_year):
    query_month = query_month.encode('ascii','ignore')
    query_year = int(query_year.encode('ascii','ignore'))
    month_year = MonthYear(query_month, query_year)
    program = 'NRLP'
    wb = Workbook()
    
    #Delete default sheet
    sh = wb.get_sheet_by_name('Sheet')
    wb.remove_sheet(sh)
    
    """Next 3 lines can be put up in for loop if there are multiple programs"""
    ws = wb.create_sheet(0)
    state_results = query_db(month_year, wb, ws, program)
    wb, ws = insert_data_excel(wb, ws, state_results, month_year, program)
    
    return wb

def query_db(month_year, wb, ws, program):    
    states = State.objects.all()
    
    state_results = {}
    for state in states:
        state_result = {}
        
        """Get current month's progress"""
        try:
            state_result['Progress'] = Progress.objects.get(state=state, project__project_name__iexact=program, year=month_year.year, month=month_year.month)
        except Progress.DoesNotExist:
            state_result['Progress'] = None
        
        """Get annual target's data"""
        financial_year = month_year.get_financial_year()
        try:
            state_result['Target'] = Target.objects.get(state=state, project__project_name__iexact=program, year=financial_year)
        except Target.DoesNotExist:
            state_result['Target'] = None
        
        """Progress till March 2013 for calculating cumulative progress"""
        try:
            state_result['TillMar13'] = ProgressTill13.objects.get(state=state, project__project_name__iexact=program, year=2013, month='Mar')
        except ProgressTill13.DoesNotExist:
            state_result['TillMar13'] = None

        "Progress till August 2013 for calculating cumulative progress"        
        try:
            state_result['TillAug13'] = ProgressTill13.objects.get(state=state, project__project_name__iexact=program, year=2013, month='Aug')
        except ProgressTill13.DoesNotExist:
            state_result['TillAug13'] = None
  
        """Add progress till last FY here"""        
        all_months = month_year.get_all_months();
        prev_years = month_year.get_previous_years()
        if not month_year.is_financial_year_2013_14():
            state_result['LastFy1'] = Progress.objects.filter(state=state, project__project_name__iexact=program, year=2013, month__in=['Sep', 'Oct', 'Nov', 'Dec'])
            if month_year.is_financial_year_2014_15():
                state_result['LastFy2'] = Progress.objects.filter(state=state, project__project_name__iexact=program, year=2014, month__in=['Jan', 'Feb', 'Mar'])
            else:
                state_result['LastFy2'] = Progress.objects.filter(state=state, project__project_name__iexact=program, year__in=prev_years, month__in=all_months)
                state_result['LastFy3'] = Progress.objects.filter(state=state, project__project_name__iexact=program, year__in=month_year.get_financial_year()-1, month__in=['Jan', 'Feb', 'Mar'])

        """Progress till last month"""        
        previous_months = month_year.get_previous_months()
        if previous_months == []:
            state_result['PrevMonth_length'] = 0
        else:
            state_result['PrevMonth_length'] = len(previous_months)
            num=1
            for year, month_list in previous_months.iteritems():
                state_result['PrevMonth'+str(num)] = Progress.objects.filter(state=state, project__project_name__iexact=program, year=year, month__in=month_list)
                num = num + 1
        
        state_results[state.state_name] = state_result
    return state_results

def insert_data_excel(wb, ws, state_results, month_year, program):
    ws.title = program
    wb, ws = design_headings_excel(wb, ws)
    row_num = 2
    target_column = 2
    for state in state_results:
        
        #Populate States
        wb, ws = populate_state_excel(wb, ws, row_num, state)
        
        #Populate Target
        target = state_results[state]['Target']
        wb, ws = populate_progess_target(wb, ws, row_num, 2, target, 0)

        #Populate current progress
        progress = state_results[state]['Progress']
        wb, ws = populate_progess_target(wb, ws, row_num, 4, progress, 1)
        
        #Populate progress upto previous month
        if month_year.is_financial_year_2013_14():
            progress_till_aug13 = state_results[state]['TillAug13'] 
            if month_year.month_index == 9:
                wb, ws = populate_progess_target(wb, ws, row_num, 3, progress_till_aug13, 1)
            elif state_results[state]['PrevMonth_length'] == 1:
                last_month_progress = state_results[state]['PrevMonth1']
                wb, ws = populate_multi_table_progress(wb, ws, row_num, 3, progress_till_aug13, last_month_progress, None, None)
            else:
                last_month_progress1 = state_results[state]['PrevMonth1']
                last_month_progress2 = state_results[state]['PrevMonth2']
                wb, ws = populate_multi_table_progress(wb, ws, row_num, 3, progress_till_aug13, last_month_progress1, last_month_progress2, None)
        elif state_results[state]['PrevMonth_length'] == 0:
            wb, ws = populate_multi_table_progress(wb, ws, row_num, 3, None, None, None, None)
        elif state_results[state]['PrevMonth_length'] == 1:
            last_month_progress = state_results[state]['PrevMonth1']
            wb, ws = populate_multi_table_progress(wb, ws, row_num, 3, None, last_month_progress, None, None)
        else:
            last_month_progress1 = state_results[state]['PrevMonth1']
            last_month_progress2 = state_results[state]['PrevMonth2']
            wb, ws = populate_multi_table_progress(wb, ws, row_num, 3, None, last_month_progress1, last_month_progress2, None)
            
        #Populate progress upto last FY
        fy_progress_2013 = state_results[state]['TillMar13']
        if month_year.is_financial_year_2013_14:
            wb, ws = populate_progess_target(wb, ws, row_num, 1, fy_progress_2013, 1)
        elif month_year.is_financial_year_2014_15:
            lastfy1 = state_results[state]['LastFy1']
            lastfy2 = state_results[state]['LastFy2']
            wb, ws = populate_multi_table_progress(wb, ws, row_num, 1, fy_progress, lastfy1, lastfy2, None)
        else:
            lastfy1 = state_results[state]['LastFy1']
            lastfy2 = state_results[state]['LastFy2']
            lastfy3 = state_results[state]['LastFy3']
            wb, ws = populate_multi_table_progress(wb, ws, row_num, 1, fy_progress, lastfy1, lastfy2, lastfy3)
        
        wb, ws = populate_cummulative(wb, ws, row_num, 5)
        row_num = row_num + 1
    return wb, ws

def populate_cummulative(wb, ws, row, col_start):
    for i in range(col_start,132,5):
            c = ws.cell(row=row, column=i)
            c.value = ws.cell(row=row, column=i-4).value + ws.cell(row=row, column=i-2).value + ws.cell(row=row, column=i-1).value
    return wb, ws
    
def populate_multi_table_progress(wb, ws, row, col_start, non_iterable_object1, table2, table3, table4):
    if not non_iterable_object1 and not table2 and not table3 and not table4:
        for i in range(col_start,132,5):
            c = ws.cell(row=row, column=i)
            c.value = 0
        return wb, ws
    Two_1 = 0
    Two_2 = 0
    Two_3 = 0
    Three_1 = 0 
    Three_2 = 0
    Three_3 = 0
    Three_4 = 0
    Three_5 = 0
    Three_6 = 0
    Three_7 = 0
    Three_8 = 0
    Five_1 = 0
    Five_2 = 0
    Five_3 = 0
    Five_4 = 0
    Six_1 = 0
    Six_2 = 0
    Six_3 = 0
    Six_4 = 0
    Seven_1 = 0
    Seven_3 = 0
    Seven_5 = 0
    Seven_6 = 0
    Seven_7 = 0
    Seven_8 = 0
    Seven_9 = 0

    if non_iterable_object1:
        Two_1 += non_iterable_object1.Two_1
        Two_2 += non_iterable_object1.Two_2
        Two_3 += non_iterable_object1.Two_3
        Three_1 += non_iterable_object1.Three_1
        Three_2 += non_iterable_object1.Three_2
        Three_3 += Three_1 + Three_2
        Three_4 += non_iterable_object1.Three_4
        Three_5 += non_iterable_object1.Three_5
        Three_6 += non_iterable_object1.Three_6
        Three_7 += non_iterable_object1.Three_7
        Three_8 += non_iterable_object1.Three_8
        Five_1 += non_iterable_object1.Five_1
        Five_2 += non_iterable_object1.Five_2
        Five_3 += non_iterable_object1.Five_5 + non_iterable_object1.Five_6 + non_iterable_object1.Five_7 + non_iterable_object1.Five_8
        Five_4 += non_iterable_object1.Five_10 + non_iterable_object1.Five_11 + non_iterable_object1.Five_12 + non_iterable_object1.Five_13
        Six_1 += non_iterable_object1.Six_1
        Six_2 += non_iterable_object1.Six_2
        Six_3 += non_iterable_object1.Six_5 + non_iterable_object1.Six_6 + non_iterable_object1.Six_7 + non_iterable_object1.Six_8
        Six_4 += non_iterable_object1.Six_10 + non_iterable_object1.Six_11 + non_iterable_object1.Six_12 + non_iterable_object1.Six_13
        Seven_1 += non_iterable_object1.Seven_1
        Seven_3 += non_iterable_object1.Seven_3
        Seven_5 += non_iterable_object1.Seven_5
        Seven_6 += non_iterable_object1.Seven_6
        Seven_7 += non_iterable_object1.Seven_7
        Seven_8 += non_iterable_object1.Seven_8
        Seven_9 += non_iterable_object1.Seven_9

    for table in [table2,table3,table4]:
        if table:
            for i in table:
                Two_1 += i.Two_1
                Two_2 += i.Two_2
                Two_3 += i.Two_3
                Three_1 += i.Three_1
                Three_2 += i.Three_2
                Three_3 += Three_1 + Three_2
                Three_4 += i.Three_4
                Three_5 += i.Three_5
                Three_6 += i.Three_6
                Three_7 += i.Three_7
                Three_8 += i.Three_8
                Five_1 += i.Five_1
                Five_2 += i.Five_2
                Five_3 += i.Five_5 + i.Five_6 + i.Five_7 + i.Five_8
                Five_4 += i.Five_10 + i.Five_11 + i.Five_12 + i.Five_13
                Six_1 += i.Six_1
                Six_2 += i.Six_2
                Six_3 += i.Six_5 + i.Six_6 + i.Six_7 + i.Six_8
                Six_4 += i.Six_10 + i.Six_11 + i.Six_12 + i.Six_13
                Seven_1 += i.Seven_1
                Seven_3 += i.Seven_3
                Seven_5 += i.Seven_5
                Seven_6 += i.Seven_6
                Seven_7 += i.Seven_7
                Seven_8 += i.Seven_8
                Seven_9 += i.Seven_9
    
    c = ws.cell(row=row, column=col_start)
    c.value = Two_1
    col_start = col_start + 5
    
    c = ws.cell(row=row, column=col_start)
    c.value = Two_2
    col_start = col_start + 5
    
    c = ws.cell(row=row, column=col_start)
    c.value = Two_3
    col_start = col_start + 5
    
    c = ws.cell(row=row, column=col_start)
    c.value = Three_1
    col_start = col_start + 5
    
    c = ws.cell(row=row, column=col_start)
    c.value = Three_2
    col_start = col_start + 5
    
    c = ws.cell(row=row, column=col_start)
    c.value = Three_1 + Three_2
    col_start = col_start + 5
    
    c = ws.cell(row=row, column=col_start)
    c.value = Three_4
    col_start = col_start + 5
    
    c = ws.cell(row=row, column=col_start)
    c.value = Three_5
    col_start = col_start + 5
    
    c = ws.cell(row=row, column=col_start)
    c.value = Three_6
    col_start = col_start + 5
    
    c = ws.cell(row=row, column=col_start)
    c.value = Three_7
    col_start = col_start + 5
    
    c = ws.cell(row=row, column=col_start)
    c.value = Three_8
    col_start = col_start + 5
    
    c = ws.cell(row=row, column=col_start)
    c.value = Five_1
    col_start = col_start + 5

    c = ws.cell(row=row, column=col_start)
    c.value = Five_2
    col_start = col_start + 5
    
    c = ws.cell(row=row, column=col_start)
    c.value = Five_3
    col_start = col_start + 5
    
    c = ws.cell(row=row, column=col_start)
    c.value = Five_4
    col_start = col_start + 5
    
    c = ws.cell(row=row, column=col_start)
    c.value = Six_1
    col_start = col_start + 5

    c = ws.cell(row=row, column=col_start)
    c.value = Six_2
    col_start = col_start + 5
    
    c = ws.cell(row=row, column=col_start)
    c.value = Six_3
    col_start = col_start + 5
    
    c = ws.cell(row=row, column=col_start)
    c.value = Six_4
    col_start = col_start + 5
    
    c = ws.cell(row=row, column=col_start)
    c.value = Seven_1
    col_start = col_start + 5
    
    c = ws.cell(row=row, column=col_start)
    c.value = Seven_3
    col_start = col_start + 5

    c = ws.cell(row=row, column=col_start)
    c.value = Seven_5
    col_start = col_start + 5

    c = ws.cell(row=row, column=col_start)
    c.value = Seven_6
    col_start = col_start + 5

    c = ws.cell(row=row, column=col_start)
    c.value = Seven_7
    col_start = col_start + 5
    
    c = ws.cell(row=row, column=col_start)
    c.value = Seven_8
    col_start = col_start + 5

    c = ws.cell(row=row, column=col_start)
    c.value = Seven_9
    col_start = col_start + 5
    
    return wb, ws

def populate_state_excel(wb, ws, row, state_name):
    c = ws.cell(row=row, column=0)
    c.value = state_name
    return wb, ws

def populate_progess_target(wb, ws, row, col_start, table, progress_true):
    if table == None:
        for i in range(col_start,132,5):
            c = ws.cell(row=row, column=i)
            c.value = 0
        return wb, ws
    c = ws.cell(row=row, column=col_start)
    c.value = table.Two_1
    col_start = col_start + 5
    
    c = ws.cell(row=row, column=col_start)
    c.value = table.Two_2
    col_start = col_start + 5
    
    c = ws.cell(row=row, column=col_start)
    c.value = table.Two_3
    col_start = col_start + 5
    
    c = ws.cell(row=row, column=col_start)
    c.value = table.Three_1
    col_start = col_start + 5
    
    c = ws.cell(row=row, column=col_start)
    c.value = table.Three_2
    col_start = col_start + 5
    
    c = ws.cell(row=row, column=col_start)
    c.value = table.Three_1 + table.Three_2
    col_start = col_start + 5
    
    c = ws.cell(row=row, column=col_start)
    c.value = table.Three_4
    col_start = col_start + 5
    
    c = ws.cell(row=row, column=col_start)
    c.value = table.Three_5
    col_start = col_start + 5
    
    c = ws.cell(row=row, column=col_start)
    c.value = table.Three_6
    col_start = col_start + 5
    
    c = ws.cell(row=row, column=col_start)
    c.value = table.Three_7
    col_start = col_start + 5
    
    c = ws.cell(row=row, column=col_start)
    c.value = table.Three_8
    col_start = col_start + 5
    
    if progress_true == 0:
        c = ws.cell(row=row, column=col_start)
        c.value = 0
        col_start = col_start + 5
    
        c = ws.cell(row=row, column=col_start)
        c.value = 0
        col_start = col_start + 5
    
    else:
        c = ws.cell(row=row, column=col_start)
        c.value = table.Five_1
        col_start = col_start + 5
    
        c = ws.cell(row=row, column=col_start)
        c.value = table.Five_2
        col_start = col_start + 5
    
    c = ws.cell(row=row, column=col_start)
    c.value = table.Five_5 + table.Five_6 + table.Five_7 + table.Five_8
    col_start = col_start + 5
    
    c = ws.cell(row=row, column=col_start)
    c.value = table.Five_10 + table.Five_11 + table.Five_12 + table.Five_13
    col_start = col_start + 5
    
    if progress_true == 0:
        c = ws.cell(row=row, column=col_start)
        c.value = 0
        col_start = col_start + 5
    else:
        c = ws.cell(row=row, column=col_start)
        c.value = table.Six_1
        col_start = col_start + 5

    c = ws.cell(row=row, column=col_start)
    c.value = table.Six_2
    col_start = col_start + 5
    
    c = ws.cell(row=row, column=col_start)
    c.value = table.Six_5 + table.Six_6 + table.Six_7 + table.Six_8
    col_start = col_start + 5
    
    c = ws.cell(row=row, column=col_start)
    c.value = table.Six_10 + table.Six_11 + table.Six_12 + table.Six_13
    col_start = col_start + 5
    
    c = ws.cell(row=row, column=col_start)
    c.value = table.Seven_1
    col_start = col_start + 5
    
    c = ws.cell(row=row, column=col_start)
    c.value = table.Seven_3
    col_start = col_start + 5

    c = ws.cell(row=row, column=col_start)
    c.value = table.Seven_5
    col_start = col_start + 5

    c = ws.cell(row=row, column=col_start)
    c.value = table.Seven_6
    col_start = col_start + 5

    c = ws.cell(row=row, column=col_start)
    c.value = table.Seven_7
    col_start = col_start + 5
    
    c = ws.cell(row=row, column=col_start)
    c.value = table.Seven_8
    col_start = col_start + 5

    c = ws.cell(row=row, column=col_start)
    c.value = table.Seven_9
    col_start = col_start + 5

    return wb, ws

def design_headings_excel(wb,ws):
    
    #TODO: "Make headings of target", "progress upto last FY", and "progress upto previous month since" generic

    c = ws.cell('A1')
    c.value = 'State'
    ws.merge_cells('A1:A2')
    c.style.alignment.wrap_text = True
    c.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
    c.style.alignment.vertical = Alignment.VERTICAL_CENTER
    c.style.borders.bottom.border_style = Border.BORDER_THIN
        
    #ws.merge_cells(start_row=5,start_column=5,end_row=5,end_column=10)
        
    c = ws.cell('B1')
    c.value = 'Total number of Blocks in which intensive strategy implementation is in progress'
    ws.row_dimensions[1].height = 35
    #ws.row_dimensions[2].height = 56
    ws.merge_cells('B1:F1')
    
    c = ws.cell('G1')
    c.value= 'Total number of Gram Panchayats in which intensive strategy implementation is in progress'
    ws.merge_cells('G1:K1')

    
    c = ws.cell('L1')
    c.value = 'Total number of villages in which intensive strategy implementation is in progress'
    ws.merge_cells('L1:P1')
    
    c = ws.cell('Q1')
    c.value = 'Number of New SHGs promoted with NRLM target households'
    ws.merge_cells('Q1:U1')
    
    c = ws.cell('V1')
    c.value = 'Number of Pre-NRLM SHGs brought into the NRLM fold after revival/training and strengthening'
    ws.merge_cells('V1:Z1')
    
    c = ws.cell('AA1')
    c.value = 'Total number of SHGs under NRLM fold in Intensive blocks (3.1 + 3.2)'
    ws.merge_cells('AA1:AE1')
    
    c = ws.cell('AF1')
    c.value = 'Number of SHGs provided basic SHG training at community level'
    ws.merge_cells('AF1:AJ1')
    
    c = ws.cell('AK1')
    c.value = 'Number of SHGs in which standard bookkeeping practices introduced'
    ws.merge_cells('AK1:AO1')
    
    c = ws.cell('AP1')
    c.value = 'Number of all SHGs following Pancha Sutras'
    ws.merge_cells('AP1:AT1')
    
    c = ws.cell('AU1')
    c.value = 'Number of SHGs bookkeepers identified and positioned after initial training'
    ws.merge_cells('AU1:AY1')
    
    c = ws.cell('AZ1')
    c.value = 'Number of internal CRPs identified and trained in the intensive blocks'
    ws.merge_cells('AZ1:BD1')
    
    c = ws.cell('BE1')
    c.value = 'Number of SHGs which are more than 3 month old'
    ws.merge_cells('BE1:BI1')
    
    c = ws.cell('BJ1')
    c.value = 'Number of 3 month old SHGs having bank accounts'
    ws.merge_cells('BJ1:BN1')
    
    c = ws.cell('BO1')
    c.value = 'Number of all 3 month old SHGs provided RF (5.5+5.6+5.7+5.8)'
    ws.merge_cells('BO1:BS1')
    
    c = ws.cell('BT1')
    c.value = 'Amount of RF provided to all SHGs  (Rs. In Lakhs) (5.10+5.11+5.12+5.13)'
    ws.merge_cells('BT1:BX1')
    
    c = ws.cell('BY1')
    c.value = 'Number of 6 month old SHGs in intensive blocks'
    ws.merge_cells('BY1:CC1')
    
    c = ws.cell('CD1')
    c.value = 'Number of SHGs which have prepared Micro Investment Plan (MIP)/Micro Credit Plan (MCP)'
    ws.merge_cells('CD1:CH1')
    
    c = ws.cell('CI1')
    c.value = 'Total Number of all SHGs provided CIF/VRF (6.5+6.6+6.7+6.8)'
    ws.merge_cells('CI1:CM1')
        
    c = ws.cell('CN1')
    c.value = 'Amount of CIF/VRF provided to all SHGs (in Rs.Lakhs) (6.10+6.11+6.12+6.13)'
    ws.merge_cells('CN1:CR1')
    
    c = ws.cell('CS1')
    c.value = 'Number of VOs formed'
    ws.merge_cells('CS1:CW1')
    
    c = ws.cell('CX1')
    c.value = 'Number of VOs provided training on basic VO management'
    ws.merge_cells('CX1:DB1')
    
    c = ws.cell('DC1')
    c.value = 'Number of VOs provided VRF (CIF)'
    ws.merge_cells('DC1:DG1')
    
    c = ws.cell('DH1')
    c.value = 'Amount of VRF(CIF) provided to VOs (in Rs.Lakhs)'
    ws.merge_cells('DH1:DL1')
    
    c = ws.cell('DM1')
    c.value = 'Number of CLFs formed'
    ws.merge_cells('DM1:DQ1')
    
    c = ws.cell('DR1')
    c.value = 'Number of CLFs provided CIF'
    ws.merge_cells('DR1:DV1')
    
    c = ws.cell('DW1')
    c.value = 'Amount of CIF provided to CLFs (in Rs.Lakhs)'
    ws.merge_cells('DW1:EA1')
    
    for i in xrange(1,131,5):
        c = ws.cell(row=0, column=i)
        c.style.alignment.wrap_text = True
        c.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
        c.style.alignment.vertical = Alignment.VERTICAL_CENTER
        c.style.borders.left.border_style = Border.BORDER_THIN
        c.style.borders.bottom.border_style = Border.BORDER_THIN
        for j in range(1,5):
            c = ws.cell(row=0, column=i+j)
            c.style.borders.bottom.border_style = Border.BORDER_THIN
    c.style.borders.right.border_style = Border.BORDER_THIN        
    
    for i in xrange(1,131,5):
        for j in range(0,5):
            c = ws.cell(row = 1, column = i+j)
            c.style.alignment.wrap_text = True
            c.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
            c.style.alignment.vertical = Alignment.VERTICAL_CENTER
            c.style.borders.left.border_style = Border.BORDER_THIN
            c.style.borders.bottom.border_style = Border.BORDER_THIN
            c.style.borders.top.border_style = Border.BORDER_THIN
            c.style.font.size = 8
            if j == 0:
                c.value = "Progress upto Mar'13"
            elif j == 1:
                c.value = "Annual Target \nFY 2013-14"
            elif j == 2:
                c.value = "Progress upto previous month since Apr'13"
            elif j == 3:
                c.value = "Progress during the Reporting Month"
            else:
                c.value = "Cummulative progress (since inception of NRLM)"
    c.style.borders.right.border_style = Border.BORDER_THIN
    return wb, ws